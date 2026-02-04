import csv
import io
import json
import os
from dataclasses import dataclass, field
from typing import Callable, Dict, List, Tuple

from dotenv import load_dotenv
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook

from app.modules.graph_store import GraphStore
from app.modules.llm_handler import LLMHandler
from app.modules.query_agent import QueryAgent
from app.modules.vector_store import SimpleVectorStore

from app.modules.file_utils import log_json, log_text


load_dotenv()


@dataclass(frozen=True)
class DomainConfig:
    key: str
    label: str
    required_columns: List[str]
    numeric_columns: Dict[str, Callable[[object], object]]
    rag_role: str
    ontology_text: str
    canonical_patterns: str
    forbidden_patterns: str
    domain_guidance: str
    question_placeholder: str
    faq: List[Dict[str, str]]


SUPPLY_CHAIN_CONFIG = DomainConfig(
    key="supplychain",
    label="Supply Chain Assistant",
    required_columns=[
        "product_id",
        "product_name",
        "component_id",
        "component_name",
        "supplier_id",
        "supplier_name",
        "supplier_country",
        "factory_id",
        "factory_name",
        "factory_country",
        "port_id",
        "port_name",
        "port_country",
        "market_country",
        "ship_cost_usd",
        "ship_time_days",
        "ship_co2_kg",
        "export_cost_usd",
        "export_time_days",
        "export_co2_kg",
        "import_cost_usd",
        "import_time_days",
        "import_co2_kg",
    ],
    numeric_columns={
        "ship_cost_usd": float,
        "ship_time_days": int,
        "ship_co2_kg": float,
        "export_cost_usd": float,
        "export_time_days": int,
        "export_co2_kg": float,
        "import_cost_usd": float,
        "import_time_days": int,
        "import_co2_kg": float,
    },
    rag_role="supply chain analyst",
    ontology_text=GraphStore.get_supply_chain_ontology_text(),
    canonical_patterns=(
        "- USES:        (p:Product)-[:USES]->(comp:Component)\n"
        "- SUPPLIES:    (s:Supplier)-[:SUPPLIES]->(comp:Component)\n"
        "- PRODUCES:    (f:Factory)-[:PRODUCES]->(p:Product)\n"
        "- LOCATED_IN:  (x)-[:LOCATED_IN]->(c:Country)\n"
        "- SHIPS_TO:    (s:Supplier)-[:SHIPS_TO]->(f:Factory)\n"
        "- EXPORTS_VIA: (f:Factory)-[:EXPORTS_VIA]->(port:Port)\n"
        "- IMPORTS_TO:  (port:Port)-[:IMPORTS_TO]->(c:Country)\n"
    ),
    forbidden_patterns=(
        "- (:Component)-[:SUPPLIES]->(:Supplier)\n"
        "- (:Product)-[:PRODUCES]->(:Factory)\n"
        "- (:Country)-[:LOCATED_IN]->(:Supplier)\n"
    ),
    domain_guidance="Supply chain questions focus on dependencies, costs, time, emissions, and disruption risk.",
    question_placeholder="Ask a question about the supply chain...",
    faq=[
        {
            "label": "Supplier B outage impact",
            "question": "What happens if Supplier B is offline for two weeks?",
        },
        {
            "label": "Carbon tax sensitivity",
            "question": "How would a $0.10/kg CO₂ carbon tax change total cost by product?",
        },
        {
            "label": "Lead time cap risk",
            "question": "Which products break first if max lead time is capped at 10 days?",
        },
        {
            "label": "Vietnam disruption",
            "question": "Why does a disruption in Vietnam affect Product Gamma but not Product Delta?",
        },
    ],
)

FRAUD_CONFIG = DomainConfig(
    key="fraudfinder",
    label="Fraud Finder",
    required_columns=[
        "account_id",
        "account_name",
        "account_status",
        "account_risk_score",
        "transaction_id",
        "transaction_amount",
        "transaction_timestamp",
        "transaction_channel",
        "merchant_id",
        "merchant_name",
        "merchant_category",
        "device_id",
        "device_type",
        "device_fingerprint",
        "email",
        "phone",
        "ip_address",
        "billing_address",
        "billing_country",
    ],
    numeric_columns={
        "transaction_amount": float,
        "account_risk_score": float,
    },
    rag_role="fraud analyst",
    ontology_text=(
        "Nodes:\n"
        "- Account {id, name, status, risk_score}\n"
        "- Transaction {id, amount, timestamp, channel}\n"
        "- Merchant {id, name, category}\n"
        "- Device {id, type, fingerprint}\n"
        "- Email {id, address}\n"
        "- Phone {id, number}\n"
        "- IP {id, address}\n"
        "- Address {id, country}\n\n"
        "Relationships:\n"
        "- (Account)-[:INITIATED]->(Transaction)\n"
        "- (Transaction)-[:TO]->(Merchant)\n"
        "- (Account)-[:USES_DEVICE]->(Device)\n"
        "- (Account)-[:USES_EMAIL]->(Email)\n"
        "- (Account)-[:USES_PHONE]->(Phone)\n"
        "- (Account)-[:LOGGED_IN_FROM]->(IP)\n"
        "- (Account)-[:HAS_ADDRESS]->(Address)\n\n"
        "IDs:\n"
        "- account_id, transaction_id, merchant_id, device_id are node ids.\n"
    ),
    canonical_patterns=(
        "- INITIATED:   (a:Account)-[:INITIATED]->(t:Transaction)\n"
        "- TO:          (t:Transaction)-[:TO]->(m:Merchant)\n"
        "- USES_DEVICE: (a:Account)-[:USES_DEVICE]->(d:Device)\n"
        "- USES_EMAIL:  (a:Account)-[:USES_EMAIL]->(e:Email)\n"
        "- USES_PHONE:  (a:Account)-[:USES_PHONE]->(p:Phone)\n"
        "- LOGGED_IN:   (a:Account)-[:LOGGED_IN_FROM]->(ip:IP)\n"
        "- HAS_ADDRESS: (a:Account)-[:HAS_ADDRESS]->(addr:Address)\n"
    ),
    forbidden_patterns="",
    domain_guidance="Fraud questions focus on shared entities, suspicious connectivity, and anomalous transaction patterns.",
    question_placeholder="Ask a question about fraud risk...",
    faq=[
        {
            "label": "General check",
            "question": "is there anything fishy going on here, based on the data?",
        },
        {
            "label": "Ring detection",
            "question": "Show accounts connected through the same merchant and device.",
        },
        {
            "label": "IP clustering",
            "question": "Which accounts log in from the same IP as flagged accounts?",
        },
        {
            "label": "Account explanation",
            "question": "Why was Account A17 flagged as high risk?",
        },
    ],
)

MEDICAL_CONFIG = DomainConfig(
    key="drhouse",
    label="Dr House",
    required_columns=[
        "patient_id",
        "patient_name",
        "visit_id",
        "visit_date",
        "provider_id",
        "provider_name",
        "symptom",
        "diagnosis",
        "medication",
        "lab_name",
        "lab_value",
        "lab_unit",
    ],
    numeric_columns={
        "lab_value": float,
    },
    rag_role="clinical decision-support assistant",
    ontology_text=(
        "Nodes:\n"
        "- Patient {id, name}\n"
        "- Visit {id, date}\n"
        "- Provider {id, name}\n"
        "- Symptom {id, name}\n"
        "- Diagnosis {id, name}\n"
        "- Medication {id, name}\n"
        "- LabTest {id, name, value, unit}\n\n"
        "Relationships:\n"
        "- (Patient)-[:HAD_VISIT]->(Visit)\n"
        "- (Provider)-[:ATTENDED]->(Visit)\n"
        "- (Visit)-[:HAS_SYMPTOM]->(Symptom)\n"
        "- (Visit)-[:RESULTED_IN]->(Diagnosis)\n"
        "- (Visit)-[:PRESCRIBED]->(Medication)\n"
        "- (Visit)-[:HAD_LAB]->(LabTest)\n\n"
        "IDs:\n"
        "- patient_id, visit_id, provider_id are node ids.\n"
    ),
    canonical_patterns=(
        "- HAD_VISIT:   (p:Patient)-[:HAD_VISIT]->(v:Visit)\n"
        "- ATTENDED:    (prov:Provider)-[:ATTENDED]->(v:Visit)\n"
        "- HAS_SYMPTOM: (v:Visit)-[:HAS_SYMPTOM]->(s:Symptom)\n"
        "- RESULTED_IN: (v:Visit)-[:RESULTED_IN]->(d:Diagnosis)\n"
        "- PRESCRIBED:  (v:Visit)-[:PRESCRIBED]->(m:Medication)\n"
        "- HAD_LAB:     (v:Visit)-[:HAD_LAB]->(l:LabTest)\n"
    ),
    forbidden_patterns="",
    domain_guidance="Medical questions should summarize evidence from visits, symptoms, labs, and diagnoses without claiming certainty.",
    question_placeholder="Ask a question about patient patterns...",
    faq=[
        {
            "label": "Symptom clusters",
            "question": "Which symptoms co-occur most often for Patient P12?",
        },
        {
            "label": "Diagnosis rationale",
            "question": "What evidence supports the diagnosis for Patient P12?",
        },
        {
            "label": "Medication patterns",
            "question": "Which medications are most common for this diagnosis?",
        },
        {
            "label": "Lab anomalies",
            "question": "Show abnormal lab patterns for Patient P12.",
        },
    ],
)

DOMAINS: Dict[str, DomainConfig] = {
    SUPPLY_CHAIN_CONFIG.key: SUPPLY_CHAIN_CONFIG,
    FRAUD_CONFIG.key: FRAUD_CONFIG,
    MEDICAL_CONFIG.key: MEDICAL_CONFIG,
}

NOTES_DIR = "notes"
DEV_PERSIST_DB_ENV = "DEV_PERSIST_DB"


@dataclass
class LLMConfig:
    provider: str
    model: str
    embed_model: str


@dataclass
class WorldState:
    rows: List[Dict] = field(default_factory=list)
    notes: Dict[int, str] = field(default_factory=dict)
    vector_store: SimpleVectorStore = field(default_factory=SimpleVectorStore)
    last_scenario: Dict = field(default_factory=dict)
    llm_config: LLMConfig | None = None
    graph_store: GraphStore | None = None
    def reset(self) -> None:
        self.rows = []
        self.notes = {}
        self.vector_store = SimpleVectorStore()
        self.last_scenario = {}

DOMAIN_STATES: Dict[str, WorldState] = {}

app = FastAPI()
app.mount("/static", StaticFiles(directory="static"), name="static")


def is_truthy(value: str | None) -> bool:
    if value is None:
        return False
    return value.strip().lower() in {"1", "true", "yes", "on"}


def dev_persist_enabled() -> bool:
    return is_truthy(os.getenv(DEV_PERSIST_DB_ENV))


def normalize_domain(domain: str | None) -> str:
    if not domain:
        return SUPPLY_CHAIN_CONFIG.key
    key = domain.strip().lower().replace(" ", "_")
    return key if key in DOMAINS else SUPPLY_CHAIN_CONFIG.key


def get_domain_config(domain: str | None) -> DomainConfig:
    return DOMAINS[normalize_domain(domain)]


def get_state(domain: str | None) -> WorldState:
    key = normalize_domain(domain)
    if key not in DOMAIN_STATES:
        DOMAIN_STATES[key] = WorldState()
    return DOMAIN_STATES[key]


def load_notes_from_disk(domain: str | None, state: WorldState) -> None:
    if state.notes:
        return
    domain_key = normalize_domain(domain)
    notes_dir = os.path.join(NOTES_DIR, domain_key)
    if not os.path.isdir(notes_dir):
        return
    notes: Dict[int, str] = {}
    for filename in sorted(os.listdir(notes_dir)):
        if not filename.startswith("note_row_") or not filename.endswith(".txt"):
            continue
        try:
            note_id = int(filename.replace("note_row_", "").replace(".txt", ""))
        except ValueError:
            continue
        path = os.path.join(notes_dir, filename)
        with open(path, "r", encoding="utf-8") as handle:
            notes[note_id] = handle.read()
    state.notes = notes


def ensure_vector_store(llm: LLMHandler, domain: str | None, state: WorldState) -> None:
    if state.vector_store.docs:
        return
    if not state.notes:
        load_notes_from_disk(domain, state)
    if state.notes:
        state.vector_store = build_vector_store(llm, state)


@app.get("/")
def index() -> HTMLResponse:
    return FileResponse("static/index.html")


@app.post("/data/upload")
def upload_data(
    file: UploadFile = File(...),
    provider: str | None = Form(default=None),
    model: str | None = Form(default=None),
    embed_model: str | None = Form(default=None),
    domain: str | None = Form(default=None),
) -> Dict:
    domain_config = get_domain_config(domain)
    state = get_state(domain_config.key)
    state.reset()
    filename = file.filename or ""
    content = file.file.read()
    try:
        if filename.lower().endswith(".csv"):
            rows = parse_csv(content, domain_config)
        elif filename.lower().endswith(".xlsx"):
            rows = parse_xlsx(content, domain_config)
        else:
            raise HTTPException(status_code=400, detail="Unsupported file type")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    state.rows = rows
    build_notes(rows, domain_config, state)

    llm_config = resolve_llm_config(provider, model, embed_model)
    state.llm_config = llm_config

    llm = build_llm(llm_config)
    try:
        state.vector_store = build_vector_store(llm, state)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Embedding ingestion failed: {exc}") from exc

    try:
        graph_store = get_graph_store(domain_config.key)
        graph_store.clear_graph()
        graph_store.insert_rows(domain_config.key, rows)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Neo4j ingestion failed: {exc}") from exc

    state.last_scenario = {}

    return {
        "status": "ok",
        "rows_loaded": len(rows),
        "schema_valid": True,
        "llm_provider": llm_config.provider,
        "llm_model": llm_config.model,
        "embed_model": llm_config.embed_model,
        "domain": domain_config.key,
    }


@app.post("/chat/rag")
def chat_rag(payload: Dict) -> Dict:
    print("RAG PAYLOAD:\n", payload)
    question = (payload.get("question") or "").strip()
    domain_config = get_domain_config(payload.get("domain"))
    state = get_state(domain_config.key)
    scenario = payload.get("scenario") or {}
    if not question:
        raise HTTPException(status_code=400, detail="Question is required")

    llm_config = resolve_llm_config(
        payload.get("provider"),
        payload.get("model"),
        payload.get("embed_model"),
    )
    llm = build_llm(llm_config)

    if not state.rows and not dev_persist_enabled():
        raise HTTPException(status_code=400, detail="No data uploaded")

    ensure_vector_store(llm, domain_config.key, state)
    retrieved = retrieve_notes(question, llm, state.vector_store)
    print("RAG RETRIEVED NOTES:\n", retrieved)
    log_json("rag_retrieved.json", {"retrieved": retrieved})
    scenario_text = scenario_summary(domain_config.key, scenario)
    print("RAG SCENARIO TEXT:\n", scenario_text)
    log_json("rag_scenario.json", {"scenario_text": scenario_text})
    table_context = ""
    if state.rows and len(state.rows) <= 200:
        table_context = rows_to_csv(state.rows, domain_config, max_rows=200)
    answer = rag_answer(
        llm,
        question,
        scenario_text,
        retrieved,
        domain_config,
        table_context or None,
    )
    print("RAG ANSWER:\n", answer)
    log_json("rag_answer.json", {"answer": answer})
    retval = {
        "answer": answer,
        "notes": retrieved,
        "retrieval": {
            "top_k": len(retrieved),
            "matches": retrieved,
        },
        "scenario": scenario_text,
        "llm_provider": llm_config.provider,
        "llm_model": llm_config.model,
        "domain": domain_config.key,
    }
    log_json("rag_response.json", retval)
    return retval


@app.post("/chat/graphrag")
def chat_graphrag(payload: Dict) -> Dict:
    print("GRAPHRAG PAYLOAD:\n", payload)
    question = (payload.get("question") or "").strip()
    domain_config = get_domain_config(payload.get("domain"))
    state = get_state(domain_config.key)
    scenario = payload.get("scenario") or {}
    if not question:
        raise HTTPException(status_code=400, detail="Question is required")

    llm_config = resolve_llm_config(
        payload.get("provider"),
        payload.get("model"),
        payload.get("embed_model"),
    )
    llm = build_llm(llm_config)

    scenario_text = scenario_summary(domain_config.key, scenario)
    graph_store = get_graph_store(domain_config.key)
    agent = QueryAgent(llm, graph_store, domain_config)

    if not state.rows and not dev_persist_enabled():
        raise HTTPException(status_code=400, detail="No data uploaded")
    if not state.rows and dev_persist_enabled() and not graph_store.has_data():
        raise HTTPException(status_code=400, detail="No graph data available")

    connection = agent.ask_cypher(question, scenario_text)

    graph = build_graph_from_results(graph_store, connection.get("results") or [])
    verification = run_verification_query(
        agent=agent,
        graph_store=graph_store,
        question=question,
        scenario_text=scenario_text,
        answer=connection.get("answer") or "",
        queries=connection.get("queries") or [],
        results=connection.get("results") or [],
    )
    retval = {
        "answer": connection.get("answer"),
        "scenario": scenario_text,
        "queries": connection.get("queries"),
        "results": connection.get("results"),
        "graph": graph,
        "verification": verification,
        "llm_provider": llm_config.provider,
        "llm_model": llm_config.model,
        "domain": domain_config.key,
    }
    print("GRAPHRAG RESPONSE:\n", retval)
    log_json("graphrag_response.json", retval)
    return retval


@app.post("/chat/rag/stream")
def chat_rag_stream(payload: Dict):
    question = (payload.get("question") or "").strip()
    domain_config = get_domain_config(payload.get("domain"))
    state = get_state(domain_config.key)
    scenario = payload.get("scenario") or {}
    if not question:
        raise HTTPException(status_code=400, detail="Question is required")

    llm_config = resolve_llm_config(
        payload.get("provider"),
        payload.get("model"),
        payload.get("embed_model"),
    )
    llm = build_llm(llm_config)

    if not state.rows and not dev_persist_enabled():
        raise HTTPException(status_code=400, detail="No data uploaded")

    ensure_vector_store(llm, domain_config.key, state)
    retrieved = retrieve_notes(question, llm, state.vector_store)
    scenario_text = scenario_summary(domain_config.key, scenario)
    table_context = ""
    if state.rows and len(state.rows) <= 200:
        table_context = rows_to_csv(state.rows, domain_config, max_rows=200)

    prompt = build_rag_prompt(
        question,
        scenario_text,
        retrieved,
        domain_config,
        table_context or None,
    )

    def event_stream():
        yield f"data: {json.dumps({'type': 'status', 'message': 'Retrieving context...'})}\n\n"
        yield f"data: {json.dumps({'type': 'meta', 'retrieval': retrieved, 'scenario': scenario_text})}\n\n"
        yield f"data: {json.dumps({'type': 'status', 'message': 'Generating answer...'})}\n\n"
        for chunk in llm.stream(prompt, temperature=0.2):
            if "error" in chunk:
                yield f"data: {json.dumps({'type': 'error', 'message': chunk['error']})}\n\n"
                return
            content = chunk.get("content", "")
            if content:
                yield f"data: {json.dumps({'type': 'token', 'content': content})}\n\n"
        yield "data: {\"type\": \"done\"}\n\n"

    return StreamingResponse(event_stream(), media_type="text/event-stream")


@app.post("/chat/graphrag/stream")
def chat_graphrag_stream(payload: Dict):
    question = (payload.get("question") or "").strip()
    domain_config = get_domain_config(payload.get("domain"))
    state = get_state(domain_config.key)
    scenario = payload.get("scenario") or {}
    if not question:
        raise HTTPException(status_code=400, detail="Question is required")

    llm_config = resolve_llm_config(
        payload.get("provider"),
        payload.get("model"),
        payload.get("embed_model"),
    )
    llm = build_llm(llm_config)

    if not state.rows and not dev_persist_enabled():
        raise HTTPException(status_code=400, detail="No data uploaded")

    scenario_text = scenario_summary(domain_config.key, scenario)
    graph_store = get_graph_store(domain_config.key)
    agent = QueryAgent(llm, graph_store, domain_config)
    if not state.rows and dev_persist_enabled() and not graph_store.has_data():
        raise HTTPException(status_code=400, detail="No graph data available")

    def event_stream():
        yield f"data: {json.dumps({'type': 'status', 'message': 'Generating Cypher queries...'})}\n\n"
        queries = agent.generate_cypher_queries_only(question, scenario_text)
        if queries:
            log_json("cypher_queries.json", {"queries": queries})
        yield f"data: {json.dumps({'type': 'status', 'message': 'Querying database...'})}\n\n"
        results = agent.execute_cypher_queries(queries) if queries else []
        log_json("cypher_execution_results.json", {"results": results})
        graph = build_graph_from_results(graph_store, results)
        payload = {
            "type": "queries",
            "queries": queries,
            "results": results,
            "scenario": scenario_text,
            "graph": graph,
        }
        yield f"data: {json.dumps(payload)}\n\n"
        yield f"data: {json.dumps({'type': 'status', 'message': 'Generating answer...'})}\n\n"
        prompt = agent.build_cypher_answer_prompt(question, scenario_text, results)
        answer_chunks = []
        for chunk in llm.stream(prompt, temperature=0.2):
            if "error" in chunk:
                yield f"data: {json.dumps({'type': 'error', 'message': chunk['error']})}\n\n"
                return
            content = chunk.get("content", "")
            if content:
                answer_chunks.append(content)
                yield f"data: {json.dumps({'type': 'token', 'content': content})}\n\n"
        answer_text = "".join(answer_chunks)
        verification = run_verification_query(
            agent=agent,
            graph_store=graph_store,
            question=question,
            scenario_text=scenario_text,
            answer=answer_text,
            queries=queries,
            results=results,
        )
        yield f"data: {json.dumps({'type': 'verify', 'verification': verification})}\n\n"
        yield "data: {\"type\": \"done\"}\n\n"

    return StreamingResponse(event_stream(), media_type="text/event-stream")


def chat_graphrag_legacy(payload: Dict) -> Dict:
    question = (payload.get("question") or "").strip()
    domain_config = get_domain_config(payload.get("domain"))
    state = get_state(domain_config.key)
    scenario = payload.get("scenario") or {}
    if not state.rows:
        raise HTTPException(status_code=400, detail="No data uploaded")
    if not question:
        raise HTTPException(status_code=400, detail="Question is required")

    llm_config = resolve_llm_config(
        payload.get("provider"),
        payload.get("model"),
        payload.get("embed_model"),
    )
    llm = build_llm(llm_config)

    scenario_text = scenario_summary(domain_config.key, scenario)
    graph_store = get_graph_store(domain_config.key)
    agent = QueryAgent(llm, graph_store, domain_config)
    connection = agent.ask(question, scenario_text)

    relationships = connection.get("relationships", [])
    trace = build_relationship_trace(relationships)
    trace_summary = f"Found {len(relationships)} graph relationships."

    if not relationships:
        schema_summary = graph_store.get_schema_summary()
        ensure_vector_store(llm, domain_config.key, state)
        retrieved = retrieve_notes(question, llm, state.vector_store)
        answer = graphrag_fallback_answer(
            llm,
            question,
            scenario_text,
            schema_summary,
            retrieved,
            domain_config,
        )
        connection["answer"] = answer
        connection["retrieval"] = {
            "top_k": len(retrieved),
            "matches": retrieved,
        }
        trace_summary = "No explicit relationships found; answered with schema + retrieved notes."

    return {
        "answer": connection.get("answer"),
        "trace": trace,
        "trace_summary": trace_summary,
        "scenario": scenario_text,
        "cypher": connection.get("cypher"),
        "params": connection.get("params"),
        "retrieval": connection.get("retrieval"),
        "llm_provider": llm_config.provider,
        "llm_model": llm_config.model,
        "domain": domain_config.key,
    }


@app.post("/chat/both")
def chat_both(payload: Dict) -> Dict:
    rag = chat_rag(payload)
    graphrag = chat_graphrag(payload)
    return {"rag": rag, "graphrag": graphrag}


def build_graph_from_results(graph_store: GraphStore, results: List[Dict]) -> Dict:
    if not results:
        return {"nodes": [], "edges": []}
    id_keys = {
        "supplier_id",
        "component_id",
        "product_id",
        "factory_id",
        "port_id",
    }
    collected = set()
    def collect_from_value(value: object) -> None:
        if value is None:
            return
        if isinstance(value, dict):
            value_type = value.get("_type")
            if value_type == "node":
                node_id = value.get("id")
                if node_id:
                    collected.add(str(node_id))
                return
            if value_type == "relationship":
                source_id = value.get("source")
                target_id = value.get("target")
                if source_id:
                    collected.add(str(source_id))
                if target_id:
                    collected.add(str(target_id))
                return
            if "id" in value:
                node_id = value.get("id")
                if node_id:
                    collected.add(str(node_id))
            for nested in value.values():
                collect_from_value(nested)
            return
        if isinstance(value, list):
            for item in value:
                collect_from_value(item)
            return
    for result in results:
        for row in result.get("rows") or []:
            if not isinstance(row, dict):
                continue
            for key, value in row.items():
                collect_from_value(value)
                if key in id_keys or key.endswith("_id"):
                    if value is None:
                        continue
                    collected.add(str(value))
    if not collected:
        return {"nodes": [], "edges": []}
    collected_list = sorted(collected)
    if len(collected_list) > 80:
        collected_list = collected_list[:80]
    return graph_store.fetch_subgraph(collected_list)


def run_verification_query(
    agent: QueryAgent,
    graph_store: GraphStore,
    question: str,
    scenario_text: str | None,
    answer: str,
    queries: List[Dict],
    results: List[Dict],
) -> Dict:
    verify_query = agent.generate_verification_query(
        question=question,
        scenario_text=scenario_text,
        answer=answer,
        queries=queries,
        results=results,
    )
    log_text("verification_query.txt", json.dumps(verify_query, indent=2, ensure_ascii=False))
    log_json("verification_queries.json", {"query": verify_query})
    if not verify_query:
        return {"query": None, "result": None, "graph": {"nodes": [], "edges": []}}
    try:
        rows = graph_store.run_cypher(verify_query["cypher"], verify_query.get("params") or {})
        result = {
            "cypher": verify_query["cypher"],
            "params": verify_query.get("params") or {},
            "reason": verify_query.get("reason"),
            "row_count": len(rows),
            "rows": rows[:50],
        }
    except Exception as exc:
        result = {
            "cypher": verify_query["cypher"],
            "params": verify_query.get("params") or {},
            "reason": verify_query.get("reason"),
            "error": str(exc),
            "row_count": 0,
            "rows": [],
        }
    log_text("verification_cypher_results.txt", json.dumps(result, indent=2, ensure_ascii=False))
    graph = build_graph_from_results(graph_store, [result])
    log_text("verification_graph.json", json.dumps(graph, indent=2, ensure_ascii=False))
    return {"query": verify_query, "result": result, "graph": graph}


def parse_csv(content: bytes, domain_config: DomainConfig) -> List[Dict]:
    text = content.decode("utf-8")
    reader = csv.DictReader(io.StringIO(text))
    return validate_rows(list(reader), domain_config)


def parse_xlsx(content: bytes, domain_config: DomainConfig) -> List[Dict]:
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("No rows found in workbook")
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    data_rows = []
    for row in rows[1:]:
        data = {headers[i]: row[i] for i in range(len(headers))}
        data_rows.append(data)
    return validate_rows(data_rows, domain_config)


def validate_rows(raw_rows: List[Dict], domain_config: DomainConfig) -> List[Dict]:
    if not raw_rows:
        raise ValueError("No data rows found")
    columns = list(raw_rows[0].keys())
    missing = [col for col in domain_config.required_columns if col not in columns]
    if missing:
        raise ValueError(f"Missing columns: {', '.join(missing)}")
    validated = []
    for index, row in enumerate(raw_rows, start=2):
        validated_row = {}
        for col in domain_config.required_columns:
            value = row.get(col)
            if value is None or str(value).strip() == "":
                raise ValueError(f"Row {index}: {col} is required")
            if col in domain_config.numeric_columns:
                try:
                    cast_value = domain_config.numeric_columns[col](value)
                except (TypeError, ValueError) as exc:
                    raise ValueError(f"Row {index}: {col} must be numeric") from exc
                validated_row[col] = cast_value
            else:
                validated_row[col] = str(value).strip()
        validated_row["row_id"] = index
        validated.append(validated_row)
    return validated


def build_supply_chain_note(row: Dict) -> str:
    return (
        f"For {row['product_name']}, the component {row['component_name']} is supplied by "
        f"{row['supplier_name']} in {row['supplier_country']}.\n"
        f"It is shipped to {row['factory_name']} at a cost of ${row['ship_cost_usd']}, "
        f"taking {row['ship_time_days']} days and emitting {row['ship_co2_kg']} kg CO₂.\n"
        f"The factory exports via {row['port_name']}, costing ${row['export_cost_usd']}, "
        f"taking {row['export_time_days']} days and emitting {row['export_co2_kg']} kg CO₂.\n"
        f"The shipment is imported to {row['market_country']} with cost "
        f"${row['import_cost_usd']}, time {row['import_time_days']} days and emissions "
        f"{row['import_co2_kg']} kg CO₂."
    )


def build_fraud_note(row: Dict) -> str:
    return (
        f"Account {row['account_name']} (id {row['account_id']}, status {row['account_status']}) "
        f"initiated transaction {row['transaction_id']} for ${row['transaction_amount']} "
        f"via {row['transaction_channel']} at {row['merchant_name']} ({row['merchant_category']}).\n"
        f"The account used device {row['device_id']} ({row['device_type']}) and logged in from "
        f"IP {row['ip_address']}. Contact: {row['email']} / {row['phone']}. "
        f"Billing address: {row['billing_address']} ({row['billing_country']})."
    )


def build_medical_note(row: Dict) -> str:
    return (
        f"Patient {row['patient_name']} (id {row['patient_id']}) visited on {row['visit_date']} "
        f"and was seen by {row['provider_name']}. Reported symptom: {row['symptom']}. "
        f"Diagnosis: {row['diagnosis']}. Medication: {row['medication']}. "
        f"Lab {row['lab_name']} = {row['lab_value']} {row['lab_unit']}."
    )


def build_notes(rows: List[Dict], domain_config: DomainConfig, state: WorldState) -> None:
    domain_key = domain_config.key
    notes_dir = os.path.join(NOTES_DIR, domain_key)
    os.makedirs(notes_dir, exist_ok=True)
    notes = {}
    for row in rows:
        if domain_key == SUPPLY_CHAIN_CONFIG.key:
            note = build_supply_chain_note(row)
        elif domain_key == FRAUD_CONFIG.key:
            note = build_fraud_note(row)
        else:
            note = build_medical_note(row)
        notes[row["row_id"]] = note
        with open(os.path.join(notes_dir, f"note_row_{row['row_id']}.txt"), "w", encoding="utf-8") as handle:
            handle.write(note)
    state.notes = notes


def rows_to_csv(rows: List[Dict], domain_config: DomainConfig, max_rows: int = 200) -> str:
    if not rows:
        return ""
    headers = [col for col in domain_config.required_columns if col in rows[0]]
    def render_cell(value: object) -> str:
        text = "" if value is None else str(value)
        if any(ch in text for ch in [",", "\"", "\n"]):
            text = text.replace("\"", "\"\"")
            return f"\"{text}\""
        return text
    lines = [",".join(headers)]
    for row in rows[:max_rows]:
        line = ",".join(render_cell(row.get(col)) for col in headers)
        lines.append(line)
    return "\n".join(lines)


def build_vector_store(llm: LLMHandler, state: WorldState) -> SimpleVectorStore:
    store = SimpleVectorStore()
    for note_id, text in state.notes.items():
        embedding = llm.embed(text)
        store.add(str(note_id), text, embedding)
    return store


def retrieve_notes(
    question: str, llm: LLMHandler, store: SimpleVectorStore, top_k: int = 5
) -> List[Dict]:
    if not question or not store.docs:
        return []
    query_embedding = llm.embed(question)
    matches = store.query(query_embedding, n=top_k)
    return [
        {
            "id": int(match["id"]),
            "score": match["score"],
            "text": match["text"],
        }
        for match in matches
    ]


def scenario_summary(domain: str, scenario: Dict) -> str:
    if normalize_domain(domain) != SUPPLY_CHAIN_CONFIG.key:
        return "No scenario constraints applied."
    parts = []
    if scenario.get("supplierBOutage"):
        parts.append("Supplier B outage is active.")
    if scenario.get("carbonTaxEnabled"):
        rate = scenario.get("carbonTaxRate", 0.1)
        parts.append(f"Carbon tax is enabled at ${rate}/kg CO₂.")
    max_days = scenario.get("maxLeadTimeDays")
    if max_days:
        parts.append(f"Max lead time is {max_days} days.")
    if not parts:
        return "No scenario constraints applied."
    return " ".join(parts)


def rag_answer(
    llm: LLMHandler,
    question: str,
    scenario_text: str,
    retrieved: List[Dict],
    domain_config: DomainConfig,
    table_context: str | None = None,
) -> str:
    prompt = build_rag_prompt(question, scenario_text, retrieved, domain_config, table_context)
    #print("RAG PROMPT:\n", prompt)
    log_text("rag_prompt.txt", prompt)
    return llm.call(prompt, temperature=0.2)


def build_rag_prompt(
    question: str,
    scenario_text: str,
    retrieved: List[Dict],
    domain_config: DomainConfig,
    table_context: str | None = None,
) -> str:
    notes_text = "\n\n".join([note["text"] for note in retrieved])
    table_block = f"Table (full data):\n{table_context}\n\n" if table_context else ""
    prompt = (
        f"You are a {domain_config.rag_role}. Use ONLY the provided context to answer. "
        "If the context does not contain the answer, say you do not have enough evidence.\n\n"
        f"Scenario: {scenario_text}\n\n"
        f"Context:\n{notes_text}\n\n"
        f"{table_block}"
        f"Question: {question}\n"
        "Answer:"
    )
    return prompt


def graphrag_fallback_answer(
    llm: LLMHandler,
    question: str,
    scenario_text: str,
    schema_summary: Dict,
    retrieved: List[Dict],
    domain_config: DomainConfig,
) -> str:
    notes_text = "\n\n".join([note["text"] for note in retrieved])
    prompt = (
        f"You are a {domain_config.label} graph assistant. Use the schema summary and retrieved notes "
        "to answer the question. If the answer is not present, explain the limitation.\n\n"
        f"Scenario: {scenario_text}\n\n"
        f"Graph Schema: {schema_summary}\n\n"
        f"Retrieved Notes:\n{notes_text}\n\n"
        f"Question: {question}\n"
        "Answer:"
    )
    print("GRAPHRAG FALLBACK PROMPT:\n", prompt)
    return llm.call(prompt, temperature=0.2)


def build_relationship_trace(relationships: List[Dict]) -> List[Dict]:
    trace = []
    for rel in relationships:
        rel_type = rel.get("rel_type", "REL")
        rel_props = rel.get("rel_props") or {}
        confidence = rel_props.get("confidence")
        source_span = rel_props.get("source_span")
        summary_parts = [rel_type]
        if confidence is not None:
            summary_parts.append(f"confidence {confidence:.2f}")
        if source_span:
            summary_parts.append(f"span: {source_span}")
        trace.append(
            {
                "rel_type": rel_type,
                "confidence": confidence,
                "source_span": source_span,
                "rel_props": rel_props,
                "summary": " | ".join(summary_parts),
            }
        )
    print("RELATIONSHIP TRACE:\n", trace)
    return trace


def resolve_llm_config(
    provider: str | None,
    model: str | None,
    embed_model: str | None,
) -> LLMConfig:
    resolved_provider = (provider or os.getenv("LLM_PROVIDER") or "openai").strip().lower()
    resolved_model = (model or os.getenv("LLM_MODEL") or default_model(resolved_provider)).strip()
    resolved_embed = (
        embed_model
        or os.getenv("LLM_EMBED_MODEL")
        or default_embed_model(resolved_provider)
    ).strip()
    return LLMConfig(
        provider=resolved_provider,
        model=resolved_model,
        embed_model=resolved_embed,
    )


def default_model(provider: str) -> str:
    return {
        "openai": "gpt-4o-mini",
        "ollama": "llama3.1",
    }.get(provider, "gpt-4o-mini")


def default_embed_model(provider: str) -> str:
    return {
        "openai": "text-embedding-3-small",
        "ollama": "nomic-embed-text",
    }.get(provider, "text-embedding-3-small")


def build_llm(config: LLMConfig) -> LLMHandler:
    try:
        return LLMHandler(
            provider=config.provider,
            model=config.model,
            embed_model=config.embed_model,
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"LLM init failed: {exc}") from exc


def get_graph_store(domain: str | None) -> GraphStore:
    domain_config = get_domain_config(domain)
    state = get_state(domain_config.key)
    if state.graph_store:
        return state.graph_store
    uri = os.getenv("NEO4J_URI")
    user = os.getenv("NEO4J_USER")
    password = os.getenv("NEO4J_PASSWORD")
    if not uri or not user or not password:
        raise HTTPException(status_code=500, detail="Missing Neo4j credentials in environment")
    database = (
        os.getenv(f"NEO4J_DB_{domain_config.key.upper()}")
        or os.getenv("NEO4J_DB_DEFAULT")
        or domain_config.key
    )
    try:
        store = GraphStore(uri, user, password, database=database, ontology_text=domain_config.ontology_text)
        with store.driver.session(database=database) as session:
            session.run("RETURN 1")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Neo4j connection failed: {exc}") from exc
    state.graph_store = store
    return store
